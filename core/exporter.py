"""
core/exporter.py — Export hasil validasi ke Excel (dengan warna) dan CSV.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

EXPORT_COLUMNS = [
    'FIRST_NAME', 'LAST_NAME', 'PARENT_ID', 'NIK', 'KK', 'ADDRESS', 'BORN_IN', 'BORN_AT',
    'STATUS', 'SKOR', 'TINGKAT_KESESUAIAN'
]

_FILL_VALID = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
_FILL_WARN  = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
_FILL_BAD   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
_FILL_HDR   = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
_THIN       = Side(style='thin', color='D0D0D0')
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def export_excel(df: pd.DataFrame, path: str) -> None:
    """Export ke Excel dengan pewarnaan baris berdasarkan TINGKAT_KESESUAIAN."""
    df = df.copy()
    for c in EXPORT_COLUMNS:
        if c not in df.columns:
            df[c] = ''
    df[EXPORT_COLUMNS].to_excel(path, index=False, engine='openpyxl')

    wb = load_workbook(path)
    ws = wb.active

    # Header styling
    for cell in ws[1]:
        cell.fill = _FILL_HDR
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 28

    # Find TINGKAT_KESESUAIAN column index
    tier_col = next(
        (i for i, c in enumerate(ws[1], 1) if c.value == 'TINGKAT_KESESUAIAN'),
        None
    )

    # Row styling
    for row_idx in range(2, ws.max_row + 1):
        tier = ws.cell(row=row_idx, column=tier_col).value if tier_col else None
        fill = _FILL_VALID if tier == 'Valid' else (_FILL_WARN if tier == 'Perlu Verifikasi' else _FILL_BAD)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    # Auto-width columns
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max_len + 4, 55)

    ws.freeze_panes = 'A2'
    wb.save(path)


def export_csv(df: pd.DataFrame, path: str) -> None:
    """Export ke CSV dengan encoding UTF-8 BOM agar kompatibel dengan Excel."""
    df = df.copy()
    for c in EXPORT_COLUMNS:
        if c not in df.columns:
            df[c] = ''
    df[EXPORT_COLUMNS].to_csv(path, index=False, encoding='utf-8-sig')
